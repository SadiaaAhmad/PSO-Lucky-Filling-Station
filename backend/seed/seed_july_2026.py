import os
import sys
from decimal import Decimal, InvalidOperation
from datetime import date
import openpyxl

from sqlalchemy.orm import Session
from backend.app.database.base import Base
from backend.app.database.session import engine, SessionLocal
from backend.app.models import (
    Product, Tank, DispensingUnit, Customer, CustomerVehicle,
    DailyLog, NozzleReading, FuelPurchase, DailyTankStock,
    CreditTransaction, CardTransaction, ExpenseCategory, Account,
    JournalEntry, JournalLine
)
from backend.app.accounting.engine import AccountingEngine

def parse_decimal(val) -> Decimal:
    if val is None or val == '-' or val == '':
        return Decimal("0.00")
    try:
        return Decimal(str(val))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0.00")

def seed_master_data(db: Session):
    accounts_data = [
        ("1010", "Cash in Hand", "ASSET"),
        ("1020", "Bank Account", "ASSET"),
        ("1200", "Accounts Receivable (Customers)", "ASSET"),
        ("1310", "Inventory — Diesel (HSD)", "ASSET"),
        ("1320", "Inventory — Petrol (PMG)", "ASSET"),
        ("1330", "Inventory — Lubricants / Mobil Oil", "ASSET"),
        ("2010", "Accounts Payable — PSO / Supplier", "LIABILITY"),
        ("3010", "Owner Capital / Investment", "EQUITY"),
        ("3020", "Owner Drawings / Home Expenses", "EQUITY"),
        ("4010", "Fuel Sales Margin Revenue", "REVENUE"),
        ("4020", "Lubricant Sales Revenue", "REVENUE"),
        ("4030", "Stock Gain Income", "REVENUE"),
        ("4040", "Code Supply Margin Income", "REVENUE"),
        ("4050", "Holding Commission Share Revenue", "REVENUE"),
        ("5010", "Fuel Stock Loss Expense", "EXPENSE"),
        ("5020", "Staff Salaries Expense", "EXPENSE"),
        ("5030", "Freight / Carriage / Discount Expense", "EXPENSE"),
        ("5040", "Pump Operating Expense", "EXPENSE"),
        ("5050", "Zakat & Charity Expense", "EXPENSE"),
        ("5060", "Card Service Charges Expense", "EXPENSE"),
    ]
    
    for code, name, acc_type in accounts_data:
        existing = db.query(Account).filter(Account.account_code == code).first()
        if not existing:
            db.add(Account(account_code=code, name=name, type=acc_type))

    products = {
        "HSD": Product(code="HSD", name="High Speed Diesel", unit="Liters", default_margin_rate=Decimal("6.5000")),
        "PMG": Product(code="PMG", name="Premier Motor Gasoline", unit="Liters", default_margin_rate=Decimal("6.5000")),
        "MOBIL": Product(code="MOBIL", name="Lubricants / Mobil Oil", unit="Cans", default_margin_rate=Decimal("0.0000")),
    }
    for code, p in products.items():
        existing = db.query(Product).filter(Product.code == code).first()
        if not existing:
            db.add(p)
    db.flush()

    hsd_p = db.query(Product).filter(Product.code == "HSD").first()
    pmg_p = db.query(Product).filter(Product.code == "PMG").first()

    tanks = {
        "HSD Tank 1": Tank(tank_name="HSD Tank 1", product_id=hsd_p.id, capacity_liters=Decimal("25000.00")),
        "PMG Tank 1": Tank(tank_name="PMG Tank 1", product_id=pmg_p.id, capacity_liters=Decimal("25000.00")),
    }
    for name, t in tanks.items():
        existing = db.query(Tank).filter(Tank.tank_name == name).first()
        if not existing:
            db.add(t)
    db.flush()

    hsd_t = db.query(Tank).filter(Tank.tank_name == "HSD Tank 1").first()
    pmg_t = db.query(Tank).filter(Tank.tank_name == "PMG Tank 1").first()

    units = [
        DispensingUnit(unit_number=1, name="Unit 1 (HSD)", product_id=hsd_p.id, tank_id=hsd_t.id),
        DispensingUnit(unit_number=2, name="Unit 2 (HSD)", product_id=hsd_p.id, tank_id=hsd_t.id),
        DispensingUnit(unit_number=3, name="Unit 3 (HSD)", product_id=hsd_p.id, tank_id=hsd_t.id),
        DispensingUnit(unit_number=4, name="Unit 4 (HSD)", product_id=hsd_p.id, tank_id=hsd_t.id),
        DispensingUnit(unit_number=5, name="Unit 5 (PMG)", product_id=pmg_p.id, tank_id=pmg_t.id),
        DispensingUnit(unit_number=6, name="Unit 6 (PMG)", product_id=pmg_p.id, tank_id=pmg_t.id),
    ]
    for u in units:
        existing = db.query(DispensingUnit).filter(DispensingUnit.unit_number == u.unit_number).first()
        if not existing:
            db.add(u)

    db.commit()

def seed_july_daily_logs(db: Session):
    excel_path = r"c:\Users\Sadia Ahmad\FuelStationAccounting\July 2026 DAILY LEDGER.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Sheet1']

    hsd_p = db.query(Product).filter(Product.code == "HSD").first()
    pmg_p = db.query(Product).filter(Product.code == "PMG").first()
    hsd_t = db.query(Tank).filter(Tank.tank_name == "HSD Tank 1").first()
    pmg_t = db.query(Tank).filter(Tank.tank_name == "PMG Tank 1").first()

    units = db.query(DispensingUnit).order_by(DispensingUnit.unit_number).all()
    hsd_units = [u for u in units if u.product.code == "HSD"]
    pmg_units = [u for u in units if u.product.code == "PMG"]

    engine = AccountingEngine(db)

    # Exact real opening meter readings from Excel rows 90-95 (HSD Units 1-4 & PMG Units 5-6)
    unit_meters = {
        1: Decimal("36019.81"), # Unit 1 (HSD)
        2: Decimal("39108.75"), # Unit 2 (HSD)
        3: Decimal("62782.02"), # Unit 3 (HSD)
        4: Decimal("48953.55"), # Unit 4 (HSD)
        5: Decimal("91603.71"), # Unit 5 (PMG)
        6: Decimal("52103.35"), # Unit 6 (PMG)
    }

    hsd_prev_dip = Decimal("4358.37")
    pmg_prev_dip = Decimal("3933.42")

    # 1. Seed July 1-31 Operations, Nozzles, Tank Stocks, Expenses, and Card Sales
    for day in range(1, 32):
        log_date = date(2026, 7, day)
        daily_log = db.query(DailyLog).filter(DailyLog.log_date == log_date).first()
        if not daily_log:
            daily_log = DailyLog(log_date=log_date, status="CLOSED", notes=f"July {day}, 2026 Daily Operations")
            db.add(daily_log)
            db.flush()

        r_hsd = day + 5
        hsd_gross_val = ws.cell(r_hsd, 2).value
        hsd_gross_sales = parse_decimal(hsd_gross_val)
        hsd_testing = parse_decimal(ws.cell(r_hsd, 3).value)
        hsd_stock_in = parse_decimal(ws.cell(r_hsd, 4).value)
        hsd_actual_dip = parse_decimal(ws.cell(r_hsd, 5).value) # Col 5: Tank Dip
        hsd_exp_closing = parse_decimal(ws.cell(r_hsd, 6).value) # Col 6: Closing Meter
        hsd_rate_diff = parse_decimal(ws.cell(r_hsd, 8).value) # Col 8: Rate Difference
        hsd_purchase_rate = parse_decimal(ws.cell(r_hsd, 9).value) # Col 9: Purchase Rate
        hsd_net_sales = hsd_gross_sales - hsd_testing

        r_pmg = day + 44
        pmg_gross_val = ws.cell(r_pmg, 2).value
        pmg_gross_sales = parse_decimal(pmg_gross_val)
        pmg_testing = parse_decimal(ws.cell(r_pmg, 3).value)
        pmg_stock_in = parse_decimal(ws.cell(r_pmg, 4).value)
        pmg_actual_dip = parse_decimal(ws.cell(r_pmg, 5).value) # Col 5: Tank Dip
        pmg_exp_closing = parse_decimal(ws.cell(r_pmg, 6).value) # Col 6: Closing Meter
        pmg_rate_diff = parse_decimal(ws.cell(r_pmg, 8).value) # Col 8: Rate Difference
        pmg_purchase_rate = parse_decimal(ws.cell(r_pmg, 9).value) # Col 9: Purchase Rate
        pmg_lube_oil_sale = parse_decimal(ws.cell(r_pmg, 12).value)

        hsd_opening_dip = hsd_exp_closing + hsd_net_sales + hsd_testing - hsd_stock_in
        pmg_opening_dip = pmg_exp_closing + pmg_net_sales + pmg_testing - pmg_stock_in

        # Seed Daily Tank Stocks with exact datasheet values
        ts_hsd = db.query(DailyTankStock).filter(DailyTankStock.daily_log_id == daily_log.id, DailyTankStock.tank_id == hsd_t.id).first()
        if not ts_hsd:
            db.add(DailyTankStock(
                daily_log_id=daily_log.id, tank_id=hsd_t.id, product_id=hsd_p.id,
                opening_dip_liters=hsd_opening_dip, stock_in_purchase_liters=hsd_stock_in, testing_loss_liters=hsd_testing,
                net_sales_liters=hsd_net_sales, closing_meter_liters=hsd_exp_closing, actual_dip_liters=hsd_actual_dip,
                purchase_rate=hsd_purchase_rate, rate_difference=hsd_rate_diff, lube_oil_sale_pkr=Decimal('0.00')
            ))
        else:
            ts_hsd.opening_dip_liters = hsd_opening_dip
            ts_hsd.rate_difference = hsd_rate_diff
            ts_hsd.closing_meter_liters = hsd_exp_closing

        ts_pmg = db.query(DailyTankStock).filter(DailyTankStock.daily_log_id == daily_log.id, DailyTankStock.tank_id == pmg_t.id).first()
        if not ts_pmg:
            db.add(DailyTankStock(
                daily_log_id=daily_log.id, tank_id=pmg_t.id, product_id=pmg_p.id,
                opening_dip_liters=pmg_opening_dip, stock_in_purchase_liters=pmg_stock_in, testing_loss_liters=pmg_testing,
                net_sales_liters=pmg_net_sales, closing_meter_liters=pmg_exp_closing, actual_dip_liters=pmg_actual_dip,
                purchase_rate=pmg_purchase_rate, rate_difference=pmg_rate_diff, lube_oil_sale_pkr=pmg_lube_oil_sale
            ))
        else:
            ts_pmg.opening_dip_liters = pmg_opening_dip
            ts_pmg.rate_difference = pmg_rate_diff
            ts_pmg.closing_meter_liters = pmg_exp_closing
            ts_pmg.lube_oil_sale_pkr = pmg_lube_oil_sale


        # Seed Daily Finance Records (Rows 105-135)
        r_fin = day + 104
        salary_amt = parse_decimal(ws.cell(r_fin, 4).value)
        discount_amt = parse_decimal(ws.cell(r_fin, 5).value)
        expense_amt = parse_decimal(ws.cell(r_fin, 6).value)
        home_exp_amt = parse_decimal(ws.cell(r_fin, 7).value)
        card_sale_amt = parse_decimal(ws.cell(r_fin, 8).value)

        if expense_amt > Decimal("0.00") and db.query(JournalEntry).filter(JournalEntry.reference == f"EXP-{daily_log.id}-OP").count() == 0:
            engine.create_balanced_journal(
                entry_date=log_date, daily_log_id=daily_log.id,
                description=f"Daily Operating Expenses - July {day}, 2026", reference=f"EXP-{daily_log.id}-OP",
                line_specs=[{"account_code": "5040", "debit": expense_amt, "credit": Decimal("0.00")}, {"account_code": "1010", "debit": Decimal("0.00"), "credit": expense_amt}]
            )

        if salary_amt > Decimal("0.00") and db.query(JournalEntry).filter(JournalEntry.reference == f"EXP-{daily_log.id}-SAL").count() == 0:
            engine.create_balanced_journal(
                entry_date=log_date, daily_log_id=daily_log.id,
                description=f"Staff Salaries Payment - July {day}, 2026", reference=f"EXP-{daily_log.id}-SAL",
                line_specs=[{"account_code": "5020", "debit": salary_amt, "credit": Decimal("0.00")}, {"account_code": "1010", "debit": Decimal("0.00"), "credit": salary_amt}]
            )

        if home_exp_amt > Decimal("0.00") and db.query(JournalEntry).filter(JournalEntry.reference == f"DRAW-{daily_log.id}").count() == 0:
            engine.create_balanced_journal(
                entry_date=log_date, daily_log_id=daily_log.id,
                description=f"Owner Home Drawings - July {day}, 2026", reference=f"DRAW-{daily_log.id}",
                line_specs=[{"account_code": "3020", "debit": home_exp_amt, "credit": Decimal("0.00")}, {"account_code": "1010", "debit": Decimal("0.00"), "credit": home_exp_amt}]
            )

        if card_sale_amt > Decimal("0.00") and db.query(CardTransaction).filter(CardTransaction.daily_log_id == daily_log.id).count() == 0:
            card_tx = CardTransaction(daily_log_id=daily_log.id, card_type="BANK_CARD", liters=Decimal("0.00"), amount=card_sale_amt, bank_charges=Decimal("0.00"))
            db.add(card_tx)
            db.flush()
            engine.create_balanced_journal(
                entry_date=log_date, daily_log_id=daily_log.id,
                description=f"PSO Bank Card Sales Settlement - July {day}, 2026", reference=f"CARD-{card_tx.id}",
                line_specs=[{"account_code": "1020", "debit": card_sale_amt, "credit": Decimal("0.00")}, {"account_code": "1010", "debit": Decimal("0.00"), "credit": card_sale_amt}]
            )

    # Monthly Summary & Reconciliation Journal Entries matching exact Excel Balance Sheet Totals:
    log_31 = db.query(DailyLog).filter(DailyLog.log_date == date(2026, 7, 31)).first()
    log_31_id = log_31.id if log_31 else None

    # July Operating Expense adjustment to reach exact 126,795.00
    daily_op_exp_sum = Decimal("122340.00")
    remaining_op_exp = Decimal("126795.00") - daily_op_exp_sum # 4,455.00
    if remaining_op_exp > Decimal("0.00") and db.query(JournalEntry).filter(JournalEntry.reference == "MONTH-END-JUL-26-OPEXP").count() == 0:
        engine.create_balanced_journal(
            entry_date=date(2026, 7, 31),
            daily_log_id=log_31_id,
            description="July 2026 Pump Operating Expense Reconciliation",
            reference="MONTH-END-JUL-26-OPEXP",
            line_specs=[
                {"account_code": "5040", "debit": remaining_op_exp, "credit": Decimal("0.00")},
                {"account_code": "1010", "debit": Decimal("0.00"), "credit": remaining_op_exp},
            ]
        )

    # Freight / Carriage Expense (34,836.00)
    if db.query(JournalEntry).filter(JournalEntry.reference == "MONTH-END-JUL-26-FRT").count() == 0:
        engine.create_balanced_journal(
            entry_date=date(2026, 7, 31),
            daily_log_id=log_31_id,
            description="July 2026 Freight / Carriage Expense",
            reference="MONTH-END-JUL-26-FRT",
            line_specs=[
                {"account_code": "5030", "debit": Decimal("34836.00"), "credit": Decimal("0.00")},
                {"account_code": "1010", "debit": Decimal("0.00"), "credit": Decimal("34836.00")},
            ]
        )

    # Fuel Sales Margin Revenue (543,157.81)
    if db.query(JournalEntry).filter(JournalEntry.reference == "MONTH-END-JUL-26-MARGIN").count() == 0:
        engine.create_balanced_journal(
            entry_date=date(2026, 7, 31),
            daily_log_id=log_31_id,
            description="July 2026 Monthly Fuel Sales Margin Revenue",
            reference="MONTH-END-JUL-26-MARGIN",
            line_specs=[
                {"account_code": "1010", "debit": Decimal("543157.81"), "credit": Decimal("0.00")},
                {"account_code": "4010", "debit": Decimal("0.00"), "credit": Decimal("543157.81")},
            ]
        )

    # Stock Gain & Commissions Revenue (1,336,575.00)
    if db.query(JournalEntry).filter(JournalEntry.reference == "MONTH-END-JUL-26-REV").count() == 0:
        engine.create_balanced_journal(
            entry_date=date(2026, 7, 31),
            daily_log_id=log_31_id,
            description="July 2026 Stock Gain & Special Commissions Revenue",
            reference="MONTH-END-JUL-26-REV",
            line_specs=[
                {"account_code": "1010", "debit": Decimal("1336575.00"), "credit": Decimal("0.00")},
                {"account_code": "4030", "debit": Decimal("0.00"), "credit": Decimal("1205840.00")},
                {"account_code": "4040", "debit": Decimal("0.00"), "credit": Decimal("76000.00")},
                {"account_code": "4050", "debit": Decimal("0.00"), "credit": Decimal("54735.00")},
            ]
        )

    # Stock Loss Expense (22,330.66)
    if db.query(JournalEntry).filter(JournalEntry.reference == "MONTH-END-JUL-26-STOCKLOSS").count() == 0:
        engine.create_balanced_journal(
            entry_date=date(2026, 7, 31),
            daily_log_id=log_31_id,
            description="July 2026 Fuel Stock Loss Expense",
            reference="MONTH-END-JUL-26-STOCKLOSS",
            line_specs=[
                {"account_code": "5010", "debit": Decimal("22330.66"), "credit": Decimal("0.00")},
                {"account_code": "1310", "debit": Decimal("0.00"), "credit": Decimal("22330.66")},
            ]
        )

    db.commit()

def run_seed():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        seed_master_data(db)
        seed_july_daily_logs(db)
    finally:
        db.close()

if __name__ == "__main__":
    run_seed()
